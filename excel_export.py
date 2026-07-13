"""Build the Excel workbook as an exact recreation of the CDACC master timetable —
the same 'WRITTEN ASSESSMENT' banner, session line and columns (Curriculum Cycle,
Course, Unit Code, Unit Name, Level, Duration) for EVERY master row — with two
columns added: Candidates and Room. Rows this centre sits are filled in; rows the
centre does not sit are left blank in those two columns.

A second 'Unassigned' sheet lists the centre's units that are not on the master
timetable at all."""
import io
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from allocate import group_by_session, allocate_session

HEADERS = ['Curriculum Cycle', 'Course', 'Unit Code', 'Unit Name', 'Level',
           'Duration', 'Candidates', 'Room']
NCOL = len(HEADERS)

BANNER = PatternFill('solid', fgColor='08463A')   # WRITTEN ASSESSMENT
HDR    = PatternFill('solid', fgColor='0D5B48')    # column header row
SESS   = PatternFill('solid', fgColor='E5F0EB')    # Date/Day/Session/Time line
ADD    = PatternFill('solid', fgColor='CFE7DB')    # the two added columns' header
WHITE  = Font(color='FFFFFF', bold=True)
ADDF   = Font(bold=True, color='08463A')
BOLD   = Font(bold=True)
THIN   = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR    = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _month_order(date_str):
    try: return datetime.strptime(date_str, '%d %b %Y')
    except Exception: return datetime.max


def _merge(ws, row, text, fill, font, align):
    ws.cell(row, 1).value = text
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOL)
    c = ws.cell(row, 1)
    c.fill, c.font, c.alignment = fill, font, align


def _centre_fill(data):
    """-> (msig -> (candidates, room_string)) for the rows this centre sits."""
    units, rooms = data['units'], data['rooms']
    sessions, _ = group_by_session(units)
    code_room = {}
    for k, us in sessions.items():
        used, _ = allocate_session(us, rooms)
        for room in used:
            for o in room['occupants']:
                lst = code_room.setdefault(o['code'], [])
                if room['room'] not in lst:
                    lst.append(room['room'])
    code_room = {c: ', '.join(v) for c, v in code_room.items()}
    fill = {}
    for u in units:
        sig = u.get('msig')
        if not sig:
            continue
        rm = code_room.get(u['code'], '')
        if sig in fill:
            pc, pr = fill[sig]
            rooms_ = ', '.join(dict.fromkeys(x for x in (pr, rm) if x))
            fill[sig] = (pc + u['count'], rooms_)
        else:
            fill[sig] = (u['count'], rm)
    return fill


def build_workbook(data):
    master = data.get('master', [])
    series = data.get('series', '')
    fill = _centre_fill(data)

    # group master rows by session, preserving the master's own row order
    groups = defaultdict(list)
    for m in master:
        groups[(m['date'], m['session'], m['time'], m.get('day', ''))].append(m)
    order = sorted(groups, key=lambda x: (_month_order(x[0]), x[1]))

    wb = Workbook()
    ws = wb.active; ws.title = 'Master Timetable'
    _merge(ws, 1, f"{(series + ' ') if series else ''}ASSESSMENT SERIES MASTER TIMETABLE",
           PatternFill(), Font(bold=True, size=14), CTR)
    _merge(ws, 2, f"{data['centre']}   (Centre code {data['centre_code']})   ·   "
                  f"Candidates & Room show this centre's allocation",
           PatternFill(), Font(bold=True, size=10, color='5B6B64'), CTR)
    ws.append([])

    for key in order:
        d, s, t, day = key
        _merge(ws, ws.max_row + 1, 'WRITTEN ASSESSMENT', BANNER, WHITE, CTR)
        _merge(ws, ws.max_row + 1, f"Date: {d}    Day: {day}    Session: {s}    Time: {t}",
               SESS, BOLD, LEFT)
        ws.append(HEADERS)
        for i, c in enumerate(ws[ws.max_row], 1):
            c.alignment, c.border = CTR, BORDER
            c.fill = ADD if i >= NCOL - 1 else HDR
            c.font = ADDF if i >= NCOL - 1 else WHITE
        for m in groups[key]:
            sig = (m['code'], m['level'], m['date'], m['session'], m['time'])
            cnt, rm = fill.get(sig, ('', ''))
            ws.append([m.get('cycle', ''), m.get('course', ''), m['code'], m['name'],
                       f"Level {m['level']}" if m['level'] else '',
                       m.get('duration') or '2 Hour', cnt, rm])
            for i, c in enumerate(ws[ws.max_row], 1):
                c.border = BORDER
                c.alignment = CTR if i in (5, 6, 7) else LEFT
        ws.append([])

    for i, w in enumerate([15, 26, 24, 42, 11, 11, 12, 24], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'

    # ---- centre units not on the master timetable ----
    _, unassigned = group_by_session(data['units'])
    if unassigned:
        wu = wb.create_sheet('Unassigned')
        _merge_simple(wu, 'Units not on the master timetable (no written session)')
        wu.append(['Curriculum Cycle', 'Course', 'Unit Code', 'Unit Name', 'Level', 'Candidates'])
        for c in wu[2]:
            c.font, c.fill, c.alignment, c.border = WHITE, HDR, CTR, BORDER
        for u in unassigned:
            wu.append([u.get('cycle') or '', u['course'], u['code'], u['name'],
                       f"Level {u['level']}" if u['level'] else '', u['count']])
            for c in wu[wu.max_row]:
                c.border, c.alignment = BORDER, LEFT
        for i, w in enumerate([15, 26, 24, 42, 11, 12], 1):
            wu.column_dimensions[get_column_letter(i)].width = w
        wu.auto_filter.ref = f"A2:F{wu.max_row}"
        wu.freeze_panes = 'A3'

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def _merge_simple(ws, text):
    ws.append([text]); ws['A1'].font = BOLD
    ws.merge_cells('A1:F1')


if __name__ == '__main__':
    import parsing
    regs = [parsing.pdf_bytes_to_text(open(f'JULY 2026 CYCLE {i}.pdf', 'rb').read()) for i in (1, 2, 3)]
    tt = open('JULY 2026 TT.pdf', 'rb').read()
    rooms = parsing.parse_rooms(open('ROOM ALLOCATION CDACC ASSESSMENT.docx', 'rb').read())
    data = parsing.build_data(regs, tt, rooms)
    open('mini_timetable.xlsx', 'wb').write(build_workbook(data).getvalue())
    print('wrote mini_timetable.xlsx')
